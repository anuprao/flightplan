#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 
# fp.py
#
# Copyright (c) 2017 Anup Jayapal Rao <anup.kadam@gmail.com>
#
# Ref: https://www.electricmonk.nl/docs/dependency_resolving_algorithm/dependency_resolving_algorithm.html
#

import argparse
from collections import OrderedDict 
import ctypes
import datetime
from enum import Enum, unique
import json
import logging
import math
import os
import os.path
import platform
#from queue import Queue
import shutil
import subprocess
import sys
import time
import datetime
#import threading
#from threading import Thread

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import colorama
from colorama import Fore, Back, Style

import svgwrite

calendar_dtStart = None
weekendList = None
holidayList = None
eventList = None

leaveListAll = None
leaveList_m1 = None

milestoneList = None

class weekday:
	def __init__(self, strDtStart):
		self.dtStart = datetime.datetime.fromisoformat(strDtStart)
		self.strDesc = "Weekday"

		self.bWeekend = False
		self.strDesc_Weekend = ""

		self.bHoliday = False
		self.strDesc_Holiday = ""

		self.bEvent = False
		self.strDesc_Event = ""

		self.bLeave = False
		self.strDesc_Leave = ""

		self.bMilestone = False
		self.strDesc_Milestone = ""

		self.width = None
		self.strClass = None

	def setup(self):
		pass

	def strftime(self, stFormat):
		return self.dtStart.strftime(stFormat)

	def __eq__(self, other):
			if isinstance(other, datetime.datetime):
				if self.dtStart == other:
					return True
				else:
					return False
			else:
				return False

class document:
	def __init__(self):
		self.margin = 10
		self.marginx = self.margin
		self.marginy = self.margin

		self.project_title_height = 10
		self.project_title_offy = 4

		self.ca_offx = self.marginx
		self.ca_offy = self.marginy + self.project_title_height

		self.gridx_fine = 10
		self.gridy_fine = 10

		self.gridx_reg = 50
		self.gridy_reg = 50

		self.crosshair_len_x = 10
		self.crosshair_len_y = 10
		self.crosshair_len_x_by_2 = self.crosshair_len_x /2
		self.crosshair_len_y_by_2 = self.crosshair_len_y /2

		self.Lw = 1	

		self.heightLegendBar = 30

		self.dwg = None

		self.setupDimensions(100,100)

	def setupDimensions(self, renderWidth, renderHeight):
		self.SW = renderWidth
		self.SH = renderHeight	
		self.dr_W = self.SW - 2*self.margin
		self.dr_H = self.SH - 2*self.margin - self.project_title_height	

		self.heightGrid = self.dr_H - self.heightLegendBar
		#print("self.heightGrid", self.heightGrid)	

	def prepSVG(self, fnSVG, strDocTitle, renderWidth, renderHeight):
		fnCSS = open('fp.css', 'r')
		strCSS = fnCSS.read()
		fnCSS.close()

		self.strDocTitle = strDocTitle

		self.setupDimensions(renderWidth, renderHeight)

		self.dwg = svgwrite.Drawing(fnSVG, size=(str(self.SW) + "px", str(self.SH) + "px")) # size=(800,480))
		self.dwg.defs.add(self.dwg.style(strCSS))

		self.pattern = self.dwg.defs.add(self.dwg.pattern(size=(4, 4), patternUnits="userSpaceOnUse"))
		self.pattern.add(self.dwg.circle((2, 2), 0.5, class_= "fixed_day_pattern"))	

		bkg = self.dwg.rect(insert=(0, 0), size=('100%', '100%'), fill='white')
		self.dwg.add(bkg)

	def drawGrid(self):

		#

		#

		oText = self.dwg.text(self.strDocTitle, x=[self.marginx], y=[self.marginy + self.project_title_offy], class_= "projectTitle")
		self.dwg.add(oText)

		#
		
		oRectFrame = self.dwg.rect(insert=(self.ca_offx + 0.5*self.Lw, self.ca_offy + 0.5*self.Lw), size=(str(self.dr_W) + "px", str(self.heightGrid) + "px"), rx=None, ry=None, class_= "frame")
		self.dwg.add(oRectFrame)	
		
		# Verticals
		
		for i in range(self.ca_offx + self.gridx_fine, self.ca_offx + self.dr_W, self.gridx_fine):
			Xs = i
			Ys = self.ca_offy 
			Xe = i
			Ye = self.ca_offy + self.heightGrid
			self.Lw = 1	
			
			oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe + 0.5*self.Lw, Ye - 0.5*self.Lw), class_= "grid gridFine")
			self.dwg.add(oLine)			
		
		##
		
		for i in range(self.ca_offx + self.gridx_reg, self.ca_offx + self.dr_W, self.gridx_reg):
			Xs = i
			Ys = self.ca_offy
			Xe = i
			Ye = self.ca_offy + self.crosshair_len_y_by_2
			self.Lw = 1	
			
			oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe + 0.5*self.Lw, Ye - 0.5*self.Lw), class_= "grid gridRegular")
			self.dwg.add(oLine)	
		
		
		for j in range(self.ca_offy + self.gridy_reg, self.ca_offy + self.heightGrid, self.gridy_reg):
			for i in range(self.ca_offx + self.gridx_reg, self.ca_offx + self.dr_W, self.gridx_reg):
				Xs = i
				Ys = j - self.crosshair_len_y_by_2 + 2
				Xe = i
				Ye = j + self.crosshair_len_y_by_2 - 1
				self.Lw = 1	
				
				oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe + 0.5*self.Lw, Ye - 0.5*self.Lw), class_= "grid gridRegular")
				self.dwg.add(oLine)			
		
		for i in range(self.ca_offx + self.gridx_reg, self.ca_offx + self.dr_W, self.gridx_reg):
			Xs = i
			Ys = self.ca_offy + self.heightGrid - self.crosshair_len_y_by_2
			Xe = i
			Ye = self.ca_offy + self.heightGrid
			self.Lw = 1	
			
			oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe + 0.5*self.Lw, Ye - 0.5*self.Lw), class_= "grid gridRegular")
			self.dwg.add(oLine)
					
		# Horizontals
		
		for j in range(self.ca_offy + self.gridy_fine, self.ca_offy + self.heightGrid, self.gridy_fine):
			Xs = self.ca_offx
			Ys = j
			Xe = self.ca_offx + self.dr_W
			Ye = j
			self.Lw = 1	
			
			oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe - 0.5*self.Lw, Ye + 0.5*self.Lw), class_= "grid gridFine")
			self.dwg.add(oLine)			
		
		##
		
		for j in range(self.ca_offy + self.gridy_reg, self.ca_offy + self.heightGrid, self.gridy_reg):
			Xs = self.ca_offx
			Ys = j
			Xe = self.ca_offx + self.crosshair_len_x_by_2
			Ye = j
			self.Lw = 1	
			
			oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe - 0.5*self.Lw, Ye + 0.5*self.Lw), class_= "grid gridRegular")
			self.dwg.add(oLine)	
		
		for i in range(self.ca_offx + self.gridx_reg, self.ca_offx + self.dr_W, self.gridx_reg):
			for j in range(self.ca_offy + self.gridy_reg, self.ca_offy + self.heightGrid, self.gridy_reg):
				Xs = i - self.crosshair_len_x_by_2 + 2
				Ys = j
				Xe = i + self.crosshair_len_x_by_2 - 1
				Ye = j
				self.Lw = 1	
				
				oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe - 0.5*self.Lw, Ye + 0.5*self.Lw), class_= "grid gridRegular")
				self.dwg.add(oLine)	
		
		for j in range(self.ca_offy + self.gridy_reg, self.ca_offy + self.heightGrid, self.gridy_reg):
			Xs = self.ca_offx + self.dr_W - self.crosshair_len_x_by_2
			Ys = j
			Xe = self.ca_offx + self.dr_W
			Ye = j
			self.Lw = 1	
			
			oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe - 0.5*self.Lw, Ye + 0.5*self.Lw), class_= "grid gridRegular")
			self.dwg.add(oLine)			

	def drawElements(self):
		pass

	def saveSVG(self):
		self.dwg.save()

	def renderSVG(self, fnSVG, strDocTitle):
		self.prepSVG(fnSVG, strDocTitle, 1920, 1080)
		self.drawGrid()
		self.drawElements()
		self.saveSVG()

class calendar(document):

	KEY_DATE_FORMAT = "%Y-%m-%d"

	def __init__(self, dtStart, dtEnd, holidayList, eventList, milestoneList, leaveList, trackList):
		super().__init__()
		self.dtStart = dtStart
		self.dtEnd = dtEnd

		self.holidayList = holidayList
		self.eventList = eventList
		self.milestoneList = milestoneList
		self.leaveList = leaveList
		self.trackList = trackList

		self.dtToday = datetime.datetime.today()
		
		self.dictDays = OrderedDict()

		#
		'''
		self.widthWeekDay = 50
		self.widthWeekendDay = 10
		self.widthWorkDay = 50
		self.widthHoliday = 10
		self.widthEventDay = 10
		'''

		self.widthWeekDay = 50
		self.widthWeekendDay = 50
		self.widthWorkDay = 50
		self.widthHoliday = 50
		self.widthEventDay = 50

		self.widthMilestoneMarker = 10
		self.widthTodayMarker = 1

		self.heightTrack = 30
		self.heightTask = 20
		self.heightBufferTrack = 50

		self.margin_task_x = 3
		self.margin_task_y = 4
		self.taskname_offx = 10
		self.taskname_offy = 14
		self.task_roundx = 2
		self.task_roundy = 2

		self.legendBar_offx = 0
		self.legendBar_offy = 0

		self.renderWidth = 160
		self.renderHeight = 90

	def isWeekend(self, sampleDay):
		bWeekend = False
		nWeekDay = sampleDay.dtStart.weekday()
		dayName = sampleDay.strftime(calendar.KEY_DATE_FORMAT)

		strDesc_Weekend = None
		if 5 == nWeekDay:
			bWeekend = True
			strDesc_Weekend = "Saturday of Week " + str(sampleDay.dtStart.isocalendar()[1])
			strDesc_Weekend = dayName + '\n' + strDesc_Weekend
			
		elif 6 == nWeekDay:
			bWeekend = True
			strDesc_Weekend = "Sunday of Week " + str(sampleDay.dtStart.isocalendar()[1])
			strDesc_Weekend = dayName + '\n' + strDesc_Weekend

		return (bWeekend, strDesc_Weekend)


	def isHoliday(self, dayName):
		bHoliday = False
		strDesc = None
		iterHoliday = iter(self.holidayList)
		bDone = False
		while(False == bDone):
			oHoliday = next(iterHoliday, None)
			if None == oHoliday:
				bDone = True
			else:
				if dayName == oHoliday[0] :
					strDesc = oHoliday[1]
					strDesc = dayName + '\n' + strDesc
					bHoliday = True
					bDone = True

		return (bHoliday, strDesc)

	def isEvent(self, dayName):
		bEvent = False
		strDesc = None
		iterEvent = iter(self.eventList)
		bDone = False
		while(False == bDone):
			oEventday = next(iterEvent, None)
			if None == oEventday:
				bDone = True
			else:
				if dayName == oEventday[0] :
					strDesc = oEventday[1]
					strDesc = dayName + '\n' + strDesc
					bEvent = True
					bDone = True
		
		return (bEvent, strDesc)

	def isLeave(self, dayName):
		bLeave = False
		strDesc = "Leave availed by : "
		arrMembers = []
		iterLeave = iter(self.leaveList)
		bDone = False
		while(False == bDone):
			oLeave = next(iterLeave, None)
			if None == oLeave:
				bDone = True
			else:
				if dayName == oLeave[0] :
					arrMembers.append(oLeave[1])
					bLeave = True

		strDesc = dayName + '\n' + strDesc
		strDesc = strDesc + ','.join(arrMembers)
		
		return (bLeave, strDesc)

	def isMilestone(self, dayName):
		bMilestone = False
		strDesc = "Milestone : "
		iterMilestone = iter(self.milestoneList)
		bDone = False
		while(False == bDone):
			oMilestone = next(iterMilestone, None)
			if None == oMilestone:
				bDone = True
			else:
				if dayName == oMilestone[0] :
					strDesc = strDesc + oMilestone[1]
					bMilestone = True
					bDone = True

		strDesc = dayName + '\n' + strDesc

		return (bMilestone, strDesc)

	def initDays(self, dtToday):
		tdDay = datetime.timedelta(hours=24)
		
		self.dtToday = dtToday

		self.period = self.dtEnd - self.dtStart
		self.dayAfterEnd = self.dtEnd + tdDay

		print(Fore.GREEN + 'Generating for ' + Fore.WHITE + str(self.period.days) + Fore.GREEN + ' days!')

		if 0 < self.period.days :

			tdDay = datetime.timedelta(hours=24)

			tmpDay = self.dtStart.replace()
			while self.dayAfterEnd != tmpDay:

				dayName = tmpDay.strftime(calendar.KEY_DATE_FORMAT)
				#print(dayName)

				oDay = weekday(dayName)

				retWeekend = self.isWeekend(oDay)
				if True == retWeekend[0]:
					oDay.bWeekend = True
					oDay.strDesc_Weekend = retWeekend[1]
				else:
					retHoliday = self.isHoliday(dayName)
					if True == retHoliday[0]:
						oDay.bHoliday = True
						oDay.strDesc_Holiday = retHoliday[1]
					else: 
						retEvent = self.isEvent(dayName)
						if True == retEvent[0]:
							oDay.bEvent = True
							oDay.strDesc_Event = retEvent[1]
						else: 
							retLeave = self.isLeave(dayName)
							if True == retLeave[0]:
								oDay.bLeave = True
								oDay.strDesc_Leave = retLeave[1]
							else:
								oDay.strDesc = dayName + '\n' + oDay.strDesc
				
				retMilestone = self.isMilestone(dayName)
				if True == retMilestone[0]:
					oDay.bMilestone = True
					oDay.strDesc_Milestone = retMilestone[1]

				self.dictDays[dayName] = oDay

				tmpDay = tmpDay + tdDay

	def setup(self):

		self.renderWidth = 0
		
		offx = self.ca_offx
		
		for dayName, sampleDay in self.dictDays.items():
			
			sampleDay.offx = offx
			if True == sampleDay.bWeekend :
				sampleDay.width = self.widthWeekendDay
				sampleDay.strClass = "weekend"
			else:
				if True == sampleDay.bHoliday :
					sampleDay.width = self.widthHoliday
					sampleDay.strClass = "holiday"
				else:
					if True == sampleDay.bEvent :
						sampleDay.width = self.widthEventDay
						sampleDay.strClass = "eventday"
					else:
						if True == sampleDay.bLeave :
							sampleDay.width = self.widthWeekDay
							sampleDay.strClass = "leaveday"
						else:
							sampleDay.width = self.widthWeekDay
							sampleDay.strClass = "workday"

			print(dayName, sampleDay.strDesc, sampleDay.strClass, sampleDay.strDesc_Weekend, sampleDay.strDesc_Holiday, sampleDay.strDesc_Event, sampleDay.strDesc_Leave, sampleDay.strDesc_Milestone)

			offx = offx + sampleDay.width

		#ew = self.getDayOffX(self.dtEnd)

		self.renderWidth = offx + self.ca_offx

		lastOffY = self.trackList[-1].offy 

		self.legendBar_offx = self.ca_offx
		self.legendBar_offy = self.ca_offy + lastOffY + self.heightTrack + self.heightBufferTrack

		#print('lastOffY', lastOffY)
		#print('self.heightTrack', self.heightTrack)
		#print('self.ca_offy', self.ca_offy)
		self.renderHeight = self.legendBar_offy + self.heightLegendBar + self.marginy
		#print('self.renderHeight', self.renderHeight)

	def getDayOffX(self, dtSample):

		offx = 0

		sampleDayname = dtSample.strftime(calendar.KEY_DATE_FORMAT)

		if sampleDayname in self.dictDays:

			oDayMatched = self.dictDays[sampleDayname]

			offx = oDayMatched.offx

		else:
			print(Fore.RED + 'getDayOffX : Cannot locate date, Please check range !')
			print(dtSample.isoformat())

		return offx

	def getAvailableHrsFor(self, dtSample):

		sampleDayname = dtSample.strftime(calendar.KEY_DATE_FORMAT)

		numAvailableHrs = 0

		if sampleDayname in self.dictDays:

			oDayMatched = self.dictDays[sampleDayname]

			if True == oDayMatched.bWeekend:
				numAvailableHrs = 0
			else:
				if True == oDayMatched.bHoliday:
					numAvailableHrs = 0
				else:
					if True == oDayMatched.bEvent:
						numAvailableHrs = 0
					else:
						if True == oDayMatched.bLeave:
							numAvailableHrs = 0
						else:
							numAvailableHrs = 8
		else:
			print(Fore.RED + 'getAvailableHrsFor : Cannot locate date, Please check range !')

		return numAvailableHrs

	def planEffort(self):

		bDone = False

		bSuccess = True

		errorList = []

		dtCommence = datetime.datetime.fromisoformat(self.dtStart.isoformat())
		
		for sampleTrack in self.trackList:
			
			oActivityList = sampleTrack.resolved

			tdDay = datetime.timedelta(hours=24)

			tmpDtStart = None

			tmpDtEnd = None

			tmpDtCurr = dtCommence.replace()

			for sampleTask in oActivityList:

				if False == sampleTask.bTraversed:

					#print(sampleTask.name)

					effortHrs = sampleTask.num_hrs

					numDaysToStart = sampleTask.daysToStartDate(tmpDtCurr)

					# if True == bValidStartDate:

					if 0 <=  numDaysToStart:

						#print("numDaysToStart", numDaysToStart)

						tmpDtCurr = tmpDtCurr + numDaysToStart*tdDay

						while (effortHrs > 0):

							tmpAllotedHrs = self.getAvailableHrsFor(tmpDtCurr)

							if tmpAllotedHrs > 0 :

								if None == tmpDtStart:
									tmpDtStart = tmpDtCurr

								effortHrs = effortHrs - tmpAllotedHrs

							tmpDtCurr = tmpDtCurr + tdDay

						tmpDtEnd = tmpDtCurr
						#print("tmpDtEnd", tmpDtEnd)

						bCheck = sampleTask.applyDateRange(tmpDtStart.replace(), tmpDtEnd.replace())
						if True == bCheck:
							
							tmpDtStart = None
							tmpDtEnd = None
							sampleTask.bTraversed = True

						else:

							bSuccess = False

							errorList.append("Calendar could not accommodate effort for task")
							errorList.append(sampleTask.name)
							errorList.append("on")
							errorList.append(sampleTask.dtStart.strftime(calendar.KEY_DATE_FORMAT))
							errorList.append("Suggested earliest date -->")
							errorList.append(tmpDtStart.strftime(calendar.KEY_DATE_FORMAT))

							bDone = True

					else:

						bSuccess = False

						errorList.append("Calendar could not accommodate start day for task")
						errorList.append(sampleTask.name)
						errorList.append("on")
						errorList.append(sampleTask.dtStart.strftime(calendar.KEY_DATE_FORMAT))
						errorList.append("Suggested date -->")
						errorList.append(tmpDtCurr.strftime(calendar.KEY_DATE_FORMAT))

						bDone = True

				else:

					tmpDtCurr = sampleTask.dtEnd

		return bSuccess, errorList
				
	def drawElements(self, strMember=None):
		'''
		# How to draw a line 
		Xs = self.ca_offx 
		Ys = self.ca_offy 
		Xe = self.ca_offx + 100 
		Ye = self.ca_offy + 100 + 1
		Lw = 1				
		
		oLine = self.dwg.line((Xs + 0.5*self.Lw, Ys + 0.5*self.Lw), (Xe + 0.5*self.Lw, Ye - 0.5*self.Lw), class_= "grid gridRegular")
		self.dwg.add(oLine)	
		'''

		#
		
		for sampleTrack in trackList:

			hol_offx = self.ca_offx 
			hol_offy = self.ca_offy + sampleTrack.offy
			hol_w = self.dr_W
			hol_h = self.heightTrack
			
			oTmpRect = self.dwg.rect(insert=(hol_offx + 0.5*self.Lw, hol_offy + 0.5*self.Lw), size=(hol_w, hol_h), rx=0, ry=0, class_= "track")
			self.dwg.add(oTmpRect)
		
		#

		for dayName, sampleDay in self.dictDays.items():
			hol_offx = sampleDay.offx 
			hol_offy = self.ca_offy
			hol_w = sampleDay.width
			hol_h = self.heightGrid
			
			#if "" != sampleDay.strClass:
			if True:

				oTmpRect = self.dwg.rect(insert=(hol_offx + 0.5*self.Lw, hol_offy + 0.5*self.Lw), size=(hol_w, hol_h), rx=0, ry=0, class_= sampleDay.strClass)
				
				if True == sampleDay.bWeekend:
					oTmpRect.set_desc(sampleDay.strDesc_Weekend, sampleDay.strDesc_Weekend)
				else:
					if True == sampleDay.bHoliday:
						oTmpRect.set_desc(sampleDay.strDesc_Holiday, sampleDay.strDesc_Holiday)
					else:
						if True == sampleDay.bEvent:
							oTmpRect.set_desc(sampleDay.strDesc_Event, sampleDay.strDesc_Event)
						else:
							if True == sampleDay.bLeave:
								oTmpRect.set_desc(sampleDay.strDesc_Leave, sampleDay.strDesc_Leave)
							else:
								oTmpRect.set_desc(sampleDay.strDesc, sampleDay.strDesc)


				#sampleDay.strDesc = dayName + '\n' + sampleDay.strDesc

				self.dwg.add(oTmpRect)

			if True == sampleDay.bMilestone:
				hol_offx = hol_offx - (self.widthMilestoneMarker/3) - 1
				hol_w = self.widthMilestoneMarker

				oTmpRect = self.dwg.rect(insert=(hol_offx + 0.5*self.Lw, hol_offy + 0.5*self.Lw), size=(hol_w, hol_h), rx=0, ry=0, class_= "milestone")
				oTmpRect.set_desc(sampleDay.strDesc_Milestone, sampleDay.strDesc_Milestone)
				self.dwg.add(oTmpRect)

		#

		hol_offx = self.getDayOffX(dtToday)
		hol_offy = self.ca_offy
		hol_w = self.widthTodayMarker
		hol_h = self.heightGrid
		
		oTmpRect = self.dwg.rect(insert=(hol_offx + 0.5*self.Lw, hol_offy + 0.5*self.Lw), size=(hol_w, hol_h), rx=0, ry=0, class_= "today")
		oTmpRect.set_desc("Today", "Today")
		self.dwg.add(oTmpRect)

		one_more_day = datetime.timedelta(hours=24)
		hol_offx = self.getDayOffX(dtToday + one_more_day)
		hol_offy = self.ca_offy
		hol_w = self.widthTodayMarker
		hol_h = self.heightGrid
		
		oTmpRect = self.dwg.rect(insert=(hol_offx + 0.5*self.Lw, hol_offy + 0.5*self.Lw), size=(hol_w, hol_h), rx=0, ry=0, class_= "today")
		oTmpRect.set_desc("Today", "Today")
		self.dwg.add(oTmpRect)

		#

		for sampleTrack in trackList:

			for sampleTask in sampleTrack.resolved:

				if False == sampleTask.bRendered:

					#print(sampleTask.name, end=" ")

					for depTask in sampleTask.listDependsOn:

						if sampleTrack != depTask.track:

							#print(depTask.name, end=",")

							sx = self.getDayOffX(depTask.dtEnd) - 10
							sy = self.ca_offy + depTask.track.offy + self.margin_task_y + 10
							ex = self.getDayOffX(sampleTask.dtStart) + 10
							ey = self.ca_offy + sampleTask.track.offy + self.margin_task_y + 10
							self.renderDependencyCurve(sx, sy, ex, ey)

					#print()

					#sprint(strMember, sampleTask.strMember)

					# If no specific member is mentioned, render the sampleTask
					# or,
					# If a specific member is mentioned, render the sampleTask only applicable for that member ... and not other members
					bRenderTask = False
					if None == strMember:
						bRenderTask = True
					else:
						if None != sampleTask.listMembers:
							if strMember in sampleTask.listMembers:
								bRenderTask = True
					#

					#prevTrack = None
					#prevTask = None
					#if sampleTrack :

					if True == bRenderTask:

						#print(sampleTask.name)

						rr_offx = self.getDayOffX(sampleTask.dtStart)
						rr_offy = self.ca_offy + sampleTask.track.offy  + self.margin_task_y

						rr_offx_end = self.getDayOffX(sampleTask.dtEnd)
						tw = rr_offx_end - rr_offx

						th = self.heightTask 

						tw_w_margin = tw - self.margin_task_x

						strClass = "task"
						if True == sampleTask.bHasDeps:
							strClass = strClass + " " + "depends_on"
						if True == sampleTask.bCritical:
							strClass = strClass + " " + "critical"
						if True == sampleTask.bComplete:
							strClass = strClass + " " + "complete"
						#if True == sampleTask.bFixedStartDate:
						#	strClass = strClass + " " + "fixed"		
						
						# 
						
						oTmpRect = self.dwg.rect(insert=(rr_offx + 2*self.Lw, rr_offy + 2*self.Lw), size=(tw_w_margin, th), rx=self.task_roundx, ry=self.task_roundy, class_=strClass)
						if False == sampleTask.bFixedStartDate:
							oTmpRect.set_desc(sampleTask.strDesc, sampleTask.strDesc)
						self.dwg.add(oTmpRect)

						if True == sampleTask.bFixedStartDate:
							oTmpPatternRect = self.dwg.rect(insert=(rr_offx + 2*self.Lw, rr_offy + 2*self.Lw), size=(tw_w_margin, th), rx=self.task_roundx, ry=self.task_roundy, fill=self.pattern.get_paint_server())
							oTmpPatternRect.set_desc(sampleTask.strDesc, sampleTask.strDesc)
							self.dwg.add(oTmpPatternRect)	

						en_x = rr_offx + self.taskname_offx
						en_y = rr_offy + self.taskname_offy
						oText = self.dwg.text(sampleTask.name, x=[en_x], y=[en_y], class_= "taskname")
						self.dwg.add(oText)

					sampleTask.bRendered = True

		'''
		#strCurve2 = 'M470,240 C490,290, 550,290, 570,340'
		sx = 470
		sy = 240
		ex = 570
		ey = 340
		self.renderDependencyCurve(sx, sy, ex, ey)
		'''

	def renderDependencyCurve(self, sx, sy, ex, ey):

		factor_x = 0.45
		factor_y = 0.05

		dx = (ex - sx) * factor_x
		dy = (ey - sy)* factor_y

		a1_x = sx+dx
		a1_y = sy+dy

		a2_x = ex-dx
		a2_y = ey-dy

		strCurve = "M" + str(sx) + "," + str(sy) + " C" + str(a1_x) + "," + str(a1_y) + ", " + str(a2_x) + "," + str(a2_y) + ", " + str(ex) + "," + str(ey)
		self.dwg.add(self.dwg.path( d=strCurve, stroke="#f1a441", fill="none", stroke_width=3, opacity=0.2))

	def resetTaskRenderedFlags(self):
		for sampleTrack in trackList:

			for sampleTask in sampleTrack.resolved:

				sampleTask.bRendered = False

	def renderSVG(self, fnSVG, strDocTitle, strMember=None):
		#print("self.renderWidth, self.renderHeight", self.renderWidth, self.renderHeight)
		self.prepSVG(fnSVG, strDocTitle, self.renderWidth, self.renderHeight)
		self.drawGrid()
		self.drawElements(strMember)
		self.saveSVG()

		self.resetTaskRenderedFlags()


class track:
	def __init__(self, name):
		self.name = name
		self.offy = 0
		self.resolved = []
		self.unresolved = []
		self.errorList = []

class tasknode:

	def __init__(self, name, strDesc, listMembers=None, bCritical=False, bComplete=False, track=None):
		self.name = name
		self.strDesc = strDesc
		self.listMembers = listMembers
		self.listDependsOn = []
		self.dtStart = None
		self.bFixedStartDate = False
		self.dtEnd = None
		self.num_man_hrs = 8
		self.bCritical = bCritical
		self.consider_weekend = False
		self.track = track
		self.consider_holiday = False

		self.bComplete = False

		self.bHasDeps = False
		self.bTraversed = False
		self.bRendered = False

		self.listDependencies = []

	def addDependencies(self, oTask):
		self.listDependencies.append(oTask)	

	def addDependency(self, sampleTask):
		if None == sampleTask.track:
			sampleTask.track = self.track
		self.listDependsOn.append(sampleTask)

		sampleTask.addDependencies(self)

	def setStartDate(self, strSample):
		self.dtStart = datetime.datetime.fromisoformat(strSample)
		self.bFixedStartDate = True

	def daysToStartDate(self, dtTmpCurr):
		nDays = -1

		if None == self.dtStart :
			nDays = 0

		else:
			period = self.dtStart - dtTmpCurr
			nDays = period.days	

		return nDays		

	def applyDateRange(self, dtStart, dtEnd):
		bReturn = False

		if None == self.dtStart :
			self.dtStart = dtStart
			self.dtEnd = dtEnd
			bReturn = True
		
		else:
			#print(self.dtStart.isoformat())
			#print(dtStart.isoformat(), dtEnd.isoformat())

			if dtStart.date() == self.dtStart.date():
				self.dtStart = dtStart
				self.dtEnd = dtEnd				
				bReturn = True

			if dtStart.date() < self.dtStart.date():
				self.dtStart = dtStart
				self.dtEnd = dtEnd				
				bReturn = True

		return bReturn

def dep_resolve(sampleTask, resolved, unresolved, errorList): 
	bRetVal = True
	
	unresolved.append(sampleTask)
	
	print(sampleTask.name + " sampleTask.listDependsOn " , end=":")
	for samplenode in sampleTask.listDependsOn:
		print(samplenode.name, end=':')
	print()
	
	iterTaskDep = iter(sampleTask.listDependsOn)
	
	bDone = False
	while(False == bDone):
		oSubTask = next(iterTaskDep, None)
		if None == oSubTask:
			bDone = True
		else:
			#print("samplenode")
			#for samplenode in resolved:
			#	print(samplenode.name, end=':')
				
			if oSubTask not in resolved:
				if oSubTask in unresolved:
					errorList.append("Circular reference detected ")
					errorList.append(sampleTask.name)
					errorList.append("-->")
					errorList.append(oSubTask.name)
					bRetVal = False
					bDone = True
				else:
					bRetValSubTask = dep_resolve(oSubTask, resolved, unresolved, errorList) 
					bRetVal = bRetVal and bRetValSubTask
					if False == bRetVal:
						bDone = True
	
	if True == bRetVal:
		resolved.append(sampleTask)
		unresolved.remove(sampleTask)
	
	return bRetVal

class srcinput:
	
	def checkDuplicateIds():
		pass

	def autoPrefixIds():
		pass

	def addTrack():
		pass

	def addTaskToTrack():
		pass



def days_hours_minutes(td):
	return td.days, td.seconds//3600, (td.seconds//60)%60


'''
from openpyxl import load_workbook

wb = load_workbook(filename = 'sampleinput/m1.xlsx')

ws = wb['tasksheet']

print(ws['B4'].value)
'''

if __name__ == "__main__":

	bProgramRet = False

	colorama.init(autoreset=True)

	print(Style.RESET_ALL, end="")

	#################################################################################################################################################################

	parser = argparse.ArgumentParser()

	parser.add_argument('--width', help='Width', nargs='?', const=1, type=int, default=1920)

	parser.add_argument('--force_redraw', help='Force redrawing (enable for profiling)', action='store_true')

	parser.add_argument('-i', '--input-file', help='XLSX files for parsing', nargs='+', default=[], required=True)

	args = parser.parse_args()

	print(args)

	nWidth = args.width

	bForceRedraw = args.force_redraw

	print(nWidth)

	print(bForceRedraw)

	print(args.input_file)

	#################################################################################################################################################################
	
	dtStart = datetime.datetime.fromisoformat('2020-11-12')
	dtEnd = datetime.datetime.fromisoformat('2021-01-31')
	dtToday = datetime.datetime.fromisoformat('2020-11-18')

	holidayList = [
		['2020-11-13', 'Deepavali']
	]

	eventList = [
		['2020-11-20', 'Demo Day']
	]

	milestoneList = [
		['2020-11-27', 'Tech1']
	]

	leaveList_m1 = [
		['2020-11-16', 'Member1'],
		['2020-12-01', 'Member1'],
		['2020-12-01', 'Member2']
	]

	leaveList_m2 = [
		['2020-11-24', 'Member2']
	]

	leaveListAll = []
	leaveListAll.extend(leaveList_m1)

	##

	t1 = track('t1')
	t1.offy = 50
	t2 = track('t2')
	t2.offy = 100
	trackList = []
	trackList.append(t1)
	trackList.append(t2)

	##

	a1 = tasknode('a1', 'Task A')
	a1.num_hrs = 8
	a1.track = t1

	b1 = tasknode('b1', 'Development Task B', listMembers=['Member1', 'Member2'], bCritical=True)
	b1.num_hrs = 24

	c1 = tasknode('c1', 'Task C', listMembers=['Member1'])
	c1.num_hrs = 8

	d1 = tasknode('d1', 'Task D', listMembers=['Member1'])
	d1.num_hrs = 16
	d1.bComplete = True

	e1 = tasknode('e1', 'Task E', listMembers=['Member2'])
	e1.num_hrs = 24
	e1.bComplete = True
	#e1.setStartDate('2020-11-17')
	#e1.setStartDate('2020-11-18')
	e1.setStartDate('2020-11-19')
	#e1.setStartDate('2020-11-20')

	#e1.dtStart = datetime.datetime.fromisoformat('2020-11-17')
	#e1.dtStart = datetime.datetime.fromisoformat('2020-11-18')
	#e1.dtStart = datetime.datetime.fromisoformat('2020-11-19')
	#e1.dtStart = datetime.datetime.fromisoformat('2020-11-20')

	a1.addDependency(b1)    # a depends on b
	a1.addDependency(d1)    # a depends on d
	b1.addDependency(c1)    # b depends on c
	b1.addDependency(e1)    # b depends on e
	c1.addDependency(d1)    # c depends on d
	c1.addDependency(e1)    # c depends on e
	#d.addDependency(b)

	#e1.bHasDeps = True

	##

	a2 = tasknode('a2', 'Task A')
	a2.num_hrs = 8
	a2.track = t2

	b2 = tasknode('b2', 'Development Task B', bCritical=True)
	b2.num_hrs = 24

	c2 = tasknode('c2', 'Task C')
	c2.num_hrs = 8

	d2 = tasknode('d2', 'Task D', listMembers=['Member1', 'Member2'])
	d2.num_hrs = 16

	e2 = tasknode('e2', 'Task E')
	e2.num_hrs = 24

	a2.addDependency(b2)    # a depends on b
	a2.addDependency(d2)    # a depends on d
	b2.addDependency(c2)    # b depends on c
	b2.addDependency(e2)    # b depends on e
	c2.addDependency(d2)    # c depends on d
	c2.addDependency(e2)    # c depends on e
	d2.addDependency(c1)

	#e2.bHasDeps = True

	print('-----------------------------------------------------')

	bSuccess_a1 = dep_resolve(a1, t1.resolved, t1.unresolved, t1.errorList)
	if False == bSuccess_a1:
		print(Fore.GREEN + 'Dependency resolution ' + Fore.WHITE + 'unsuccessful' + Fore.GREEN + 'Check Data!!')
		
		print('resolved :')
		for sampleTask in t1.resolved:
			print(sampleTask.name, end=':')
		print()		
		
		print('unresolved :')
		for sampleTask in t1.unresolved:
			print(sampleTask.name, end=':')
		print()

		print('errorList :')
		for item in t1.errorList:
			print(item, end='')
		print()
			
	else:
		print(Fore.GREEN + 'Dependency resolution ' + Fore.WHITE + 'successful' + Fore.GREEN + '!!!')

		for sampleTask in t1.resolved:
			print(sampleTask.name, end=':')
			#sampleTask.bTraversed = True

		print()

	print('-----------------------------------------------------')

	bSuccess_a2 = dep_resolve(a2, t2.resolved, t2.unresolved, t2.errorList)
	if False == bSuccess_a2:
		print(Fore.GREEN + 'Dependency resolution ' + Fore.WHITE + 'unsuccessful' + Fore.GREEN + 'Check Data!!')
		
		print('resolved :')
		for sampleTask in t2.resolved:
			print(sampleTask.name, end=':')
		print()		
		
		print('unresolved :')
		for sampleTask in t2.unresolved:
			print(sampleTask.name, end=':')
		print()

		print('errorList :')
		for item in t2.errorList:
			print(item, end='')
		print()
			
	else:
		print(Fore.GREEN + 'Dependency resolution ' + Fore.WHITE + 'successful' + Fore.GREEN + '!!!')

		for sampleTask in t2.resolved:
			print(sampleTask.name, end=':')
			#sampleTask.bTraversed = True

		print()

	print('-----------------------------------------------------')

	#################################################################################################################################################################
	# 	
	if bSuccess_a1 and bSuccess_a2:

		oCal = calendar(dtStart, dtEnd, holidayList, eventList, milestoneList, leaveListAll, trackList)
		oCal.initDays(dtToday)
		oCal.setup()
		
		bSuccess, planEffort_errorList = oCal.planEffort()
		if True == bSuccess:

			oCal.renderSVG('calendar_overall.svg', 'Project plan')

			oCal.renderSVG('calendar_Member1.svg', 'Project plan', 'Member1')

			oCal.renderSVG('calendar_Member2.svg', 'Project plan', 'Member2')

			bProgramRet = True

		else:
			print()
			print(Fore.RED + 'Error in planning effort :')
			for item in planEffort_errorList:
				print(item) #, end='')
			print()

			bProgramRet = False

	if False == bProgramRet:
		print(Fore.RED + 'Program did not succeed !')
	else:
		print(Fore.BLUE + 'Program completed successfully !')



